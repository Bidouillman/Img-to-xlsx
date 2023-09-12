from PIL import Image
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.utils import get_column_letter

image = 'your-img-path'
xlsname = 'your-xlxs-name' 
# Ouvrir l'image / open the image and convert it with rgb value
image = Image.open(image).convert('RGB')

# Picture dimensions 
largeur, hauteur = image.size

# New Excel 
classeur = openpyxl.Workbook()
feuille = classeur.active



for y in range(hauteur):
    for x in range(largeur):
        # extraire les valeurs RGB de chaque pixel de la ligne
        r, g, b = image.getpixel((x, y))
        
        ##separate the values to put them in 3 separate cells (one red then green then blue)
        for z in range(3):
            cellule = feuille.cell(row=y*3+z+1, column=x+1)

            if z == 0:
                value = r
                color = f"FF{r:02X}0000"
            elif z == 1:
                value = g
                color = f"FF00{g:02X}00"
            else:
                value = b
                color = f"FF0000{b:02X}"

            # Adding color and the rgb value
            cellule.value = f"{value}"
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cellule.fill = fill

        print(x, y) # show the progress

classeur.save(xlsname)
print("finish")
